"""
excel_manager.py - Excel Real-Time Manager
Setiap tambah/edit/hapus data → file Excel otomatis diperbarui
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import datetime, date
import logging
import os

from config import EXCEL_FILE

logger = logging.getLogger(__name__)

# =====================
# STYLE HELPERS
# =====================

HEADER_FILL    = PatternFill("solid", fgColor="1F4E79")   # Biru tua
MASUK_FILL     = PatternFill("solid", fgColor="E2EFDA")   # Hijau muda
KELUAR_FILL    = PatternFill("solid", fgColor="FCE4D6")   # Merah muda
SUMMARY_FILL   = PatternFill("solid", fgColor="FFF2CC")   # Kuning muda
TITLE_FILL     = PatternFill("solid", fgColor="2E75B6")   # Biru sedang

HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
TITLE_FONT   = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
BODY_FONT    = Font(name="Calibri", size=10)
TOTAL_FONT   = Font(bold=True, name="Calibri", size=11)

THIN = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

IDR_FORMAT = '#,##0'


def _apply_header(ws, row, cols, fill=HEADER_FILL, font=HEADER_FONT):
    for col, title in enumerate(cols, 1):
        c = ws.cell(row=row, column=col, value=title)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER


def _set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w


def _format_idr(val):
    return val if isinstance(val, (int, float)) else 0


# =====================
# REBUILD EXCEL PENUH
# =====================

def rebuild_excel(all_setoran, all_pengeluaran):
    """
    Bangun ulang file Excel dari nol berdasarkan data dari DB.
    Dipanggil setelah setiap operasi CRUD.
    """
    try:
        wb = openpyxl.Workbook()

        _sheet_setoran(wb, all_setoran)
        _sheet_pengeluaran(wb, all_pengeluaran)
        _sheet_summary(wb, all_setoran, all_pengeluaran)
        _sheet_harian(wb, all_setoran, all_pengeluaran)
        _sheet_mingguan(wb, all_setoran, all_pengeluaran)
        _sheet_bulanan(wb, all_setoran, all_pengeluaran)
        _sheet_tahunan(wb, all_setoran, all_pengeluaran)

        # Hapus sheet default kosong jika ada
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(EXCEL_FILE)
        logger.info(f"✅ Excel diperbarui: {EXCEL_FILE}")
    except Exception as e:
        logger.error(f"❌ Gagal update Excel: {e}")


# =====================
# SHEET 1 - SETORAN
# =====================

def _sheet_setoran(wb, data):
    ws = wb.create_sheet("💰 Setoran Masuk")
    ws.sheet_view.showGridLines = True
    ws.freeze_panes = "A3"

    # Judul
    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value = "📋 REKAP SETORAN (UANG MASUK)"
    t.font = TITLE_FONT
    t.fill = TITLE_FILL
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["ID", "Nama Penyetor", "Nominal (Rp)", "Keterangan", "Kategori",
               "Tanggal", "Waktu", "Dicatat Oleh", "Tgl Update"]
    _apply_header(ws, 2, headers)
    ws.row_dimensions[2].height = 22

    _set_col_widths(ws, [6, 22, 18, 28, 15, 14, 10, 16, 16])

    for i, d in enumerate(data, 3):
        row = [
            d['id'],
            d['nama'],
            _format_idr(d['nominal']),
            d.get('keterangan', '-'),
            d.get('kategori', 'setoran'),
            str(d['tanggal']),
            str(d['waktu'])[:8],
            d.get('username', '-'),
            str(d.get('updated_at', ''))[:19]
        ]
        for col, val in enumerate(row, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = BODY_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center")
            if col == 3:
                c.number_format = IDR_FORMAT
                c.alignment = Alignment(horizontal="right", vertical="center")
            ws.row_dimensions[i].height = 18
        # Warna zebra
        if i % 2 == 0:
            for col in range(1, 10):
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor="EBF3FB")

    # Baris total
    last = len(data) + 2
    tr = last + 1
    ws.cell(row=tr, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=tr, column=1).fill = SUMMARY_FILL
    total_cell = ws.cell(row=tr, column=3,
                         value=f"=SUM(C3:C{last})" if data else 0)
    total_cell.font = TOTAL_FONT
    total_cell.number_format = IDR_FORMAT
    total_cell.fill = SUMMARY_FILL
    total_cell.alignment = Alignment(horizontal="right")


# =====================
# SHEET 2 - PENGELUARAN
# =====================

def _sheet_pengeluaran(wb, data):
    ws = wb.create_sheet("💸 Pengeluaran")
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value = "📋 REKAP PENGELUARAN (UANG KELUAR)"
    t.font = TITLE_FONT
    t.fill = PatternFill("solid", fgColor="C00000")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["ID", "Nama / Keperluan", "Nominal (Rp)", "Keterangan", "Kategori",
               "Tanggal", "Waktu", "Dicatat Oleh", "Tgl Update"]
    _apply_header(ws, 2, headers)
    ws.row_dimensions[2].height = 22
    _set_col_widths(ws, [6, 22, 18, 28, 15, 14, 10, 16, 16])

    for i, d in enumerate(data, 3):
        row = [
            d['id'],
            d['nama'],
            _format_idr(d['nominal']),
            d.get('keterangan', '-'),
            d.get('kategori', 'pengeluaran'),
            str(d['tanggal']),
            str(d['waktu'])[:8],
            d.get('username', '-'),
            str(d.get('updated_at', ''))[:19]
        ]
        for col, val in enumerate(row, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = BODY_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center")
            if col == 3:
                c.number_format = IDR_FORMAT
                c.alignment = Alignment(horizontal="right", vertical="center")
            ws.row_dimensions[i].height = 18
        if i % 2 == 0:
            for col in range(1, 10):
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor="FDF2F2")

    last = len(data) + 2
    tr = last + 1
    ws.cell(row=tr, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=tr, column=1).fill = SUMMARY_FILL
    tc = ws.cell(row=tr, column=3, value=f"=SUM(C3:C{last})" if data else 0)
    tc.font = TOTAL_FONT
    tc.number_format = IDR_FORMAT
    tc.fill = SUMMARY_FILL
    tc.alignment = Alignment(horizontal="right")


# =====================
# SHEET 3 - SUMMARY
# =====================

def _sheet_summary(wb, setoran, pengeluaran):
    ws = wb.create_sheet("📊 Ringkasan")

    total_masuk = sum(_format_idr(d['nominal']) for d in setoran)
    total_keluar = sum(_format_idr(d['nominal']) for d in pengeluaran)
    untung_rugi = total_masuk - total_keluar

    now = datetime.now().strftime("%d/%m/%Y %H:%M")

    def row_data(ws, r, label, value, fill=None, is_idr=True):
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = Font(bold=True, name="Calibri", size=11)
        c1.border = THIN_BORDER
        if fill:
            c1.fill = fill
        c2 = ws.cell(row=r, column=2, value=value)
        c2.font = Font(bold=True, name="Calibri", size=11,
                       color="375623" if value >= 0 else "C00000" if isinstance(value, (int, float)) and value < 0 else "000000")
        c2.border = THIN_BORDER
        if fill:
            c2.fill = fill
        if is_idr and isinstance(value, (int, float)):
            c2.number_format = IDR_FORMAT
        c2.alignment = Alignment(horizontal="right")

    ws.merge_cells("A1:B1")
    t = ws["A1"]
    t.value = f"📊 RINGKASAN KESELURUHAN  —  Update: {now}"
    t.font = TITLE_FONT
    t.fill = TITLE_FILL
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22

    row_data(ws, 2, "Total Setoran Masuk",  total_masuk,   MASUK_FILL)
    row_data(ws, 3, "Total Pengeluaran",    total_keluar,  KELUAR_FILL)
    row_data(ws, 4, "UNTUNG / RUGI",        untung_rugi,
             PatternFill("solid", fgColor="FFFF00" if untung_rugi >= 0 else "FF6666"))
    row_data(ws, 5, "Jumlah Transaksi Masuk",  len(setoran),    None, False)
    row_data(ws, 6, "Jumlah Transaksi Keluar", len(pengeluaran), None, False)
    row_data(ws, 7, "Total Transaksi",
             len(setoran) + len(pengeluaran), None, False)

    for r in range(2, 8):
        ws.row_dimensions[r].height = 22


# =====================
# SHEET 4 - HARIAN
# =====================

def _sheet_harian(wb, setoran, pengeluaran):
    ws = wb.create_sheet("📅 Harian")
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "📅 REKAP HARIAN"
    t.font = TITLE_FONT
    t.fill = TITLE_FILL
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Tanggal", "Masuk (Rp)", "Keluar (Rp)", "Untung/Rugi (Rp)",
               "Jml Masuk", "Jml Keluar"]
    _apply_header(ws, 2, headers)
    _set_col_widths(ws, [16, 18, 18, 18, 12, 12])

    # Kumpulkan semua tanggal unik
    dates = sorted(set(
        [str(d['tanggal']) for d in setoran] +
        [str(d['tanggal']) for d in pengeluaran]
    ), reverse=True)

    for i, tgl in enumerate(dates, 3):
        masuk = sum(_format_idr(d['nominal']) for d in setoran if str(d['tanggal']) == tgl)
        keluar = sum(_format_idr(d['nominal']) for d in pengeluaran if str(d['tanggal']) == tgl)
        ur = masuk - keluar
        jm = sum(1 for d in setoran if str(d['tanggal']) == tgl)
        jk = sum(1 for d in pengeluaran if str(d['tanggal']) == tgl)

        row = [tgl, masuk, keluar, ur, jm, jk]
        for col, val in enumerate(row, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = BODY_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center",
                                    horizontal="right" if col > 1 else "center")
            if col in [2, 3, 4]:
                c.number_format = IDR_FORMAT
            if col == 4 and ur < 0:
                c.font = Font(bold=True, color="C00000", name="Calibri", size=10)
            elif col == 4 and ur > 0:
                c.font = Font(bold=True, color="375623", name="Calibri", size=10)
        ws.row_dimensions[i].height = 18


# =====================
# SHEET 5 - MINGGUAN
# =====================

def _sheet_mingguan(wb, setoran, pengeluaran):
    from datetime import timedelta

    ws = wb.create_sheet("📆 Mingguan")
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "📆 REKAP MINGGUAN"
    t.font = TITLE_FONT
    t.fill = TITLE_FILL
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Minggu Ke (Mulai)", "Masuk (Rp)", "Keluar (Rp)", "Untung/Rugi (Rp)",
               "Jml Masuk", "Jml Keluar"]
    _apply_header(ws, 2, headers)
    _set_col_widths(ws, [20, 18, 18, 18, 12, 12])

    # Kelompokkan per minggu (Senin awal)
    from collections import defaultdict
    week_masuk = defaultdict(int)
    week_keluar = defaultdict(int)
    week_jm = defaultdict(int)
    week_jk = defaultdict(int)

    for d in setoran:
        tgl = d['tanggal'] if isinstance(d['tanggal'], date) else date.fromisoformat(str(d['tanggal']))
        monday = tgl - timedelta(days=tgl.weekday())
        week_masuk[monday] += _format_idr(d['nominal'])
        week_jm[monday] += 1

    for d in pengeluaran:
        tgl = d['tanggal'] if isinstance(d['tanggal'], date) else date.fromisoformat(str(d['tanggal']))
        monday = tgl - timedelta(days=tgl.weekday())
        week_keluar[monday] += _format_idr(d['nominal'])
        week_jk[monday] += 1

    all_weeks = sorted(set(list(week_masuk.keys()) + list(week_keluar.keys())), reverse=True)

    for i, w in enumerate(all_weeks, 3):
        masuk = week_masuk[w]
        keluar = week_keluar[w]
        ur = masuk - keluar
        row = [str(w), masuk, keluar, ur, week_jm[w], week_jk[w]]
        for col, val in enumerate(row, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = BODY_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center",
                                    horizontal="right" if col > 1 else "center")
            if col in [2, 3, 4]:
                c.number_format = IDR_FORMAT
            if col == 4:
                c.font = Font(bold=True,
                              color="375623" if ur >= 0 else "C00000",
                              name="Calibri", size=10)
        ws.row_dimensions[i].height = 18


# =====================
# SHEET 6 - BULANAN
# =====================

def _sheet_bulanan(wb, setoran, pengeluaran):
    ws = wb.create_sheet("🗓️ Bulanan")
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "🗓️ REKAP BULANAN"
    t.font = TITLE_FONT
    t.fill = TITLE_FILL
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Bulan", "Masuk (Rp)", "Keluar (Rp)", "Untung/Rugi (Rp)",
               "Jml Masuk", "Jml Keluar"]
    _apply_header(ws, 2, headers)
    _set_col_widths(ws, [16, 18, 18, 18, 12, 12])

    from collections import defaultdict
    bln_masuk = defaultdict(int)
    bln_keluar = defaultdict(int)
    bln_jm = defaultdict(int)
    bln_jk = defaultdict(int)
    BULAN = ["", "Jan", "Feb", "Mar", "Apr", "Mei", "Jun",
             "Jul", "Agu", "Sep", "Okt", "Nov", "Des"]

    for d in setoran:
        tgl = d['tanggal'] if isinstance(d['tanggal'], date) else date.fromisoformat(str(d['tanggal']))
        key = (tgl.year, tgl.month)
        bln_masuk[key] += _format_idr(d['nominal'])
        bln_jm[key] += 1

    for d in pengeluaran:
        tgl = d['tanggal'] if isinstance(d['tanggal'], date) else date.fromisoformat(str(d['tanggal']))
        key = (tgl.year, tgl.month)
        bln_keluar[key] += _format_idr(d['nominal'])
        bln_jk[key] += 1

    all_months = sorted(set(list(bln_masuk.keys()) + list(bln_keluar.keys())), reverse=True)

    for i, (yr, mn) in enumerate(all_months, 3):
        masuk = bln_masuk[(yr, mn)]
        keluar = bln_keluar[(yr, mn)]
        ur = masuk - keluar
        label = f"{BULAN[mn]} {yr}"
        row = [label, masuk, keluar, ur, bln_jm[(yr, mn)], bln_jk[(yr, mn)]]
        for col, val in enumerate(row, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = BODY_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center",
                                    horizontal="right" if col > 1 else "center")
            if col in [2, 3, 4]:
                c.number_format = IDR_FORMAT
            if col == 4:
                c.font = Font(bold=True,
                              color="375623" if ur >= 0 else "C00000",
                              name="Calibri", size=10)
        ws.row_dimensions[i].height = 18


# =====================
# SHEET 7 - TAHUNAN
# =====================

def _sheet_tahunan(wb, setoran, pengeluaran):
    ws = wb.create_sheet("📈 Tahunan")
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "📈 REKAP TAHUNAN"
    t.font = TITLE_FONT
    t.fill = TITLE_FILL
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["Tahun", "Masuk (Rp)", "Keluar (Rp)", "Untung/Rugi (Rp)",
               "Jml Masuk", "Jml Keluar"]
    _apply_header(ws, 2, headers)
    _set_col_widths(ws, [10, 18, 18, 18, 12, 12])

    from collections import defaultdict
    thn_masuk = defaultdict(int)
    thn_keluar = defaultdict(int)
    thn_jm = defaultdict(int)
    thn_jk = defaultdict(int)

    for d in setoran:
        tgl = d['tanggal'] if isinstance(d['tanggal'], date) else date.fromisoformat(str(d['tanggal']))
        thn_masuk[tgl.year] += _format_idr(d['nominal'])
        thn_jm[tgl.year] += 1

    for d in pengeluaran:
        tgl = d['tanggal'] if isinstance(d['tanggal'], date) else date.fromisoformat(str(d['tanggal']))
        thn_keluar[tgl.year] += _format_idr(d['nominal'])
        thn_jk[tgl.year] += 1

    all_years = sorted(set(list(thn_masuk.keys()) + list(thn_keluar.keys())), reverse=True)

    for i, yr in enumerate(all_years, 3):
        masuk = thn_masuk[yr]
        keluar = thn_keluar[yr]
        ur = masuk - keluar
        row = [yr, masuk, keluar, ur, thn_jm[yr], thn_jk[yr]]
        for col, val in enumerate(row, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = BODY_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center",
                                    horizontal="right" if col > 1 else "center")
            if col in [2, 3, 4]:
                c.number_format = IDR_FORMAT
            if col == 4:
                c.font = Font(bold=True,
                              color="375623" if ur >= 0 else "C00000",
                              name="Calibri", size=10)
        ws.row_dimensions[i].height = 18
